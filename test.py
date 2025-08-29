import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
import os
import subprocess

# Étape 1 : Créer un nouveau classeur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Liste de présence"

# Étape 2 : Définir les entêtes
headers = [
    "Nom", "Prénoms", "Numéro de téléphone",
    "Contact d'urgence", "Département de service"
]

# Ajouter les dimanches d'août 2025
date = datetime(2025, 8, 1)
while date.month == 8:
    if date.weekday() == 6:  # Dimanche
        headers.append(date.strftime("%d/%m/%Y"))
    date += timedelta(days=1)

# Étape 3 : Écrire les entêtes
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

# Étape 4 : Ajouter les lignes vides pour 54 personnes
for row in range(2, 56):  # de ligne 2 à 55 inclus
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col, value="")

# Étape 5 : Enregistrer le fichier
file_path = os.path.abspath("Liste_Presence_Aout_2025.xlsx")
wb.save(file_path)

# Étape 6 : Ouvrir le fichier avec Microsoft Excel
try:
    subprocess.Popen(['start', 'excel', file_path], shell=True)
    print(f"✅ Fichier ouvert dans Microsoft Excel : {file_path}")
except Exception as e:
    print(f"❌ Impossible d’ouvrir Excel automatiquement : {e}")
